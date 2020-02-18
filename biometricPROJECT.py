import cv2
import face_recognition as fc
import random as rd
import xlrd
import smtplib as sm
import time
import datetime
from openpyxl import load_workbook
from datetime import date
def mark(i,a):
    d=str(date.today())
    k=d.split("-")
    wb=load_workbook(r"F:\Desktop\techienest\facereco.xlsx")
    ws=wb.get_active_sheet()
    if(int(k[2])==1):
        for l in range(0,31):
            ws.cell(row=i,column=(131+l)).value=None
        ws.cell(row=i,column=(130+int(k[2]))).value=a
        wb.save(r"F:\Desktop\techienest\facereco.xlsx")
sender="himanshusmack@gmail.com"
password=""
v=cv2.VideoCapture(0)
wb=xlrd.open_workbook(r"F:\Desktop\techienest\facereco.xlsx")
sheet=wb.sheet_by_index(0)
i=0
l=[]
p=[]
email=[]
name=[]
for i in range(sheet.nrows):
    for j in range(2,130):
        l.append(sheet.cell_value(i,j))
    p.append(l)
    l=[]
for i in range(sheet.nrows):
    name.append(sheet.cell_value(i,0))
    email.append(sheet.cell_value(i,1))
print(email)
while(1):
    print("WELCOME")
    print("want to mark your attendance? \n")
    ch=int(input("1.YES 2.NO"))
    now=datetime.datetime.now()
    today7pm=now.replace(hour=00,minute=00,second=0,microsecond=0)
    today730pm=now.replace(hour=21,minute=30,second=0,microsecond=0)
    if(ch==1):
        if(now>=today7pm and now<=today730pm):
            try:
                ret,k=v.read()
                f12=fc.face_locations(k)
                fc12=fc.face_encodings(k,f12)
                res=fc.compare_faces(p,fc12[0])
            except IndexError:
                print("Sorry No Face detected")
                exit(0)
            if(True in res):
                ind=res.index(True)
                print("detecting your face...")
                print("Hello"+name[ind])
                otp1=int(rd.random()*10000)
                msg="your OTP is "+str(otp1)
                s=sm.SMTP('smtp.gmail.com',587)
                s.starttls()
                s.login(sender,password)
                s.sendmail(sender,email[ind],msg)
                print("your OTP has been sent to mail")
                now1=int(time.time())
                print("your OTP will be valid for 10 min only /n")
                s.quit()
                flag=0
                for i in range(0,3):
                    otp=int(input('enter OTP'))
                    now2=int(time.time())
                    time1=now2-now1
                    if(time1<60):
                        if(otp==otp1):
                            print("you are marked presnt!")
                            mark(ind+1,'P')
                            flag=1
                            break
                        else:
                            print("incorrect otp please enter again \n")
                    else:
                        print('session expired')
                if(flag==0):
                    print("you are marked absent \n")
                    print(ind+1, 'A')
            else:
                print("your face does not match!")
                exit(0)
        else:
            print("your attendance can be marked only between 9:00 AM to 9:30 AM")
            exit(0)
    else:
        print("ok thanks!")
        exit(0)
                        
                        
